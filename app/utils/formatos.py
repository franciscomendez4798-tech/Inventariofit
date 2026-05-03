# app/utils/formatos.py
"""
Generador de formatos institucionales XLSX y PDF.

Formato R-AP-33-01-01: Requisición de Material  → para el solicitante
Formato R-AP-33-01-02: Salida de Almacén        → para el administrador (y solicitante al entregar)

Estructura del Excel original (ambos formatos):
  - 2 copias del formato en la misma hoja (filas 1-19 y 20-38)
  - 10 filas para materiales (filas 9-18 en pedido, 9-17 en salida)
  - Si hay más de 10/9 items, se generan hojas adicionales
"""
import os
import math
from copy import copy
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

# Rutas a las plantillas
_DIR = os.path.dirname(__file__)
PLANTILLA_PEDIDO = os.path.join(_DIR, 'plantilla_pedido.xlsx')
PLANTILLA_SALIDA = os.path.join(_DIR, 'plantilla_salida.xlsx')

# Directorio de salida
UPLOAD_BASE = os.path.join(os.path.dirname(_DIR), 'static', 'uploads', 'requisiciones')
UPLOAD_PEDIDOS = os.path.join(UPLOAD_BASE, 'pedidos')
UPLOAD_SALIDAS = os.path.join(UPLOAD_BASE, 'salidas')


def _copiar_estilo(celda_origen, celda_destino):
    """Copia estilos de una celda a otra."""
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = celda_origen.number_format
        celda_destino.protection = copy(celda_origen.protection)
        celda_destino.alignment = copy(celda_origen.alignment)


def _llenar_bloque_pedido(ws, fila_inicio, items, departamento, fecha, solicitante, num_hoja):
    """
    Llena un bloque de formato de pedido (19 filas).
    fila_inicio: 1 para el bloque superior, 20 para el bloque inferior (copia).
    items: lista de dicts {cantidad, unidad, descripcion}
    """
    # Fecha
    ws.cell(row=fila_inicio + 3, column=8, value=fecha.strftime('%d/%m/%Y'))

    # Departamento
    ws.cell(row=fila_inicio + 5, column=3, value=departamento)

    # Número de hoja
    ws.cell(row=fila_inicio + 5, column=8, value=num_hoja)

    # Materiales (filas 9-18 relativas → fila_inicio+8 a fila_inicio+17)
    for i, item in enumerate(items[:10]):
        fila = fila_inicio + 8 + i
        ws.cell(row=fila, column=1, value=item['cantidad'])
        ws.cell(row=fila, column=2, value=item['unidad'])
        ws.cell(row=fila, column=3, value=item['descripcion'])

    # Solicitó
    ws.cell(row=fila_inicio + 18, column=2, value=solicitante)


def _llenar_bloque_salida(ws, fila_inicio, items, departamento, fecha,
                          entrego, recibio, num_hoja):
    """
    Llena un bloque de formato de salida de almacén (19 filas).
    items: lista de dicts {cantidad, unidad, descripcion}
    """
    # Fecha
    ws.cell(row=fila_inicio + 3, column=8, value=fecha.strftime('%d/%m/%Y'))

    # Departamento
    ws.cell(row=fila_inicio + 5, column=3, value=departamento)

    # Número de hoja
    ws.cell(row=fila_inicio + 5, column=8, value=num_hoja)

    # Materiales (filas 9-17 relativas → fila_inicio+8 a fila_inicio+16)
    for i, item in enumerate(items[:9]):
        fila = fila_inicio + 8 + i
        ws.cell(row=fila, column=1, value=item['cantidad'])
        ws.cell(row=fila, column=2, value=item['unidad'])
        ws.cell(row=fila, column=3, value=item['descripcion'])

    # Entregó / Recibió
    ws.cell(row=fila_inicio + 17, column=2, value=entrego)
    ws.cell(row=fila_inicio + 18, column=2, value=recibio)


def generar_formato_pedido(pedido):
    """
    Genera el formato R-AP-33-01-01 (Requisición de Material) para un pedido.
    Retorna la ruta relativa al archivo generado.
    """
    # Preparar datos
    items = []
    for d in pedido.detalles:
        items.append({
            'cantidad': d.cantidad_solicitada,
            'unidad': d.material.unidad_medida,
            'descripcion': d.material.nombre,
        })

    departamento = pedido.departamento.nombre if pedido.departamento else 'N/D'
    fecha = pedido.fecha_solicitud or datetime.utcnow()
    solicitante = pedido.solicitante.nombre_completo if pedido.solicitante else 'N/D'

    # Items por página
    items_por_pagina = 10
    total_paginas = max(1, math.ceil(len(items) / items_por_pagina))

    # Cargar plantilla
    wb = openpyxl.load_workbook(PLANTILLA_PEDIDO)
    ws = wb.active

    # Llenar la primera página (bloque 1: filas 1-19, bloque 2: filas 20-38)
    pagina_items = items[:items_por_pagina]
    _llenar_bloque_pedido(ws, 1, pagina_items, departamento, fecha, solicitante, 1)
    _llenar_bloque_pedido(ws, 20, pagina_items, departamento, fecha, solicitante, 1)

    # Si hay más páginas, crear hojas adicionales
    for pag in range(1, total_paginas):
        inicio_items = pag * items_por_pagina
        pagina_items = items[inicio_items:inicio_items + items_por_pagina]
        num_hoja = pag + 1

        # Copiar la hoja de la plantilla
        ws_nueva = wb.copy_worksheet(wb.worksheets[0])
        ws_nueva.title = f'Hoja {num_hoja}'

        # Limpiar datos previos de la plantilla copiada
        for fila in range(9, 19):
            for col in range(1, 9):
                ws_nueva.cell(row=fila, column=col, value=None)
                ws_nueva.cell(row=fila + 19, column=col, value=None)

        _llenar_bloque_pedido(ws_nueva, 1, pagina_items, departamento, fecha,
                              solicitante, num_hoja)
        _llenar_bloque_pedido(ws_nueva, 20, pagina_items, departamento, fecha,
                              solicitante, num_hoja)

    # Guardar
    os.makedirs(UPLOAD_PEDIDOS, exist_ok=True)
    nombre = f'{pedido.folio}_requisicion.xlsx'
    ruta = os.path.join(UPLOAD_PEDIDOS, nombre)
    wb.save(ruta)
    wb.close()

    return f'uploads/requisiciones/pedidos/{nombre}'


def generar_formato_salida(pedido):
    """
    Genera el formato R-AP-33-01-02 (Salida de Almacén) para un pedido
    aprobado/entregado. Usa cantidades APROBADAS.
    Retorna la ruta relativa al archivo generado.
    """
    # Preparar datos con cantidades aprobadas
    items = []
    for d in pedido.detalles:
        cant = d.cantidad_aprobada if d.cantidad_aprobada is not None else d.cantidad_solicitada
        if cant and cant > 0:
            items.append({
                'cantidad': cant,
                'unidad': d.material.unidad_medida,
                'descripcion': d.material.nombre,
            })

    departamento = pedido.departamento.nombre if pedido.departamento else 'N/D'
    fecha = pedido.fecha_resolucion or pedido.fecha_solicitud or datetime.utcnow()
    solicitante = pedido.solicitante.nombre_completo if pedido.solicitante else 'N/D'
    # El admin que aprobó — por ahora usar texto genérico
    admin = 'Secretaría Administrativa'

    # Items por página (salida tiene 9 filas)
    items_por_pagina = 9
    total_paginas = max(1, math.ceil(len(items) / items_por_pagina))

    # Cargar plantilla
    wb = openpyxl.load_workbook(PLANTILLA_SALIDA)
    ws = wb.active

    # Primera página
    pagina_items = items[:items_por_pagina]
    _llenar_bloque_salida(ws, 1, pagina_items, departamento, fecha,
                          admin, solicitante, 1)
    _llenar_bloque_salida(ws, 20, pagina_items, departamento, fecha,
                          admin, solicitante, 1)

    # Páginas adicionales
    for pag in range(1, total_paginas):
        inicio_items = pag * items_por_pagina
        pagina_items = items[inicio_items:inicio_items + items_por_pagina]
        num_hoja = pag + 1

        ws_nueva = wb.copy_worksheet(wb.worksheets[0])
        ws_nueva.title = f'Hoja {num_hoja}'

        for fila in range(9, 18):
            for col in range(1, 9):
                ws_nueva.cell(row=fila, column=col, value=None)
                ws_nueva.cell(row=fila + 19, column=col, value=None)

        _llenar_bloque_salida(ws_nueva, 1, pagina_items, departamento, fecha,
                              admin, solicitante, num_hoja)
        _llenar_bloque_salida(ws_nueva, 20, pagina_items, departamento, fecha,
                              admin, solicitante, num_hoja)

    # Guardar
    os.makedirs(UPLOAD_SALIDAS, exist_ok=True)
    nombre = f'{pedido.folio}_salida.xlsx'
    ruta = os.path.join(UPLOAD_SALIDAS, nombre)
    wb.save(ruta)
    wb.close()

    return f'uploads/requisiciones/salidas/{nombre}'


def generar_pdf_requisiciones(pedidos, tipo='pedido'):
    """
    Genera un PDF con todos los formatos de requisición o salida
    de los pedidos dados. Cada pedido ocupa una página.

    tipo: 'pedido' → formato R-AP-33-01-01
          'salida' → formato R-AP-33-01-02

    Retorna un BytesIO con el PDF.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Spacer, Paragraph, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'FormatoTitulo', parent=styles['Heading2'],
        alignment=TA_CENTER, fontSize=12, spaceAfter=4
    )
    estilo_subtitulo = ParagraphStyle(
        'FormatoSub', parent=styles['Normal'],
        alignment=TA_CENTER, fontSize=9, textColor=colors.grey,
        spaceAfter=12
    )
    estilo_campo = ParagraphStyle(
        'Campo', parent=styles['Normal'],
        fontSize=9, spaceAfter=2
    )
    estilo_firma = ParagraphStyle(
        'Firma', parent=styles['Normal'],
        fontSize=9, spaceBefore=20
    )

    elementos = []

    for idx, pedido in enumerate(pedidos):
        if idx > 0:
            elementos.append(PageBreak())

        # Encabezado
        if tipo == 'pedido':
            titulo = 'REQUISICIÓN DE MATERIAL'
            codigo = 'R-AP-33-01-01 V4'
        else:
            titulo = 'SALIDA DE ALMACÉN'
            codigo = 'R-AP-33-01-02'

        elementos.append(Paragraph('SECRETARÍA ADMINISTRATIVA', estilo_titulo))
        elementos.append(Paragraph(titulo, estilo_titulo))
        elementos.append(Paragraph(codigo, estilo_subtitulo))

        # Info
        fecha = pedido.fecha_solicitud if tipo == 'pedido' else (
            pedido.fecha_resolucion or pedido.fecha_solicitud)
        fecha_str = fecha.strftime('%d/%m/%Y') if fecha else 'N/D'
        depto = pedido.departamento.nombre if pedido.departamento else 'N/D'

        elementos.append(Paragraph(
            f'<b>Fecha:</b> {fecha_str} &nbsp;&nbsp; '
            f'<b>Folio:</b> {pedido.folio}',
            estilo_campo
        ))
        elementos.append(Paragraph(
            f'<b>Departamento:</b> {depto}',
            estilo_campo
        ))
        elementos.append(Spacer(1, 10))

        # Tabla de materiales
        encabezados = ['Cantidad', 'Unidad', 'Descripción']
        data = [encabezados]

        for d in pedido.detalles:
            if tipo == 'pedido':
                cant = d.cantidad_solicitada
            else:
                cant = d.cantidad_aprobada if d.cantidad_aprobada is not None else d.cantidad_solicitada
            if cant and cant > 0:
                data.append([
                    str(cant),
                    d.material.unidad_medida,
                    d.material.nombre,
                ])

        col_widths = [1.2*inch, 1.2*inch, 4*inch]
        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C41E2A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F5F5F5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elementos.append(tabla)

        # Firmas
        solicitante = pedido.solicitante.nombre_completo if pedido.solicitante else 'N/D'
        if tipo == 'pedido':
            elementos.append(Paragraph(
                f'<b>Solicitó:</b> {solicitante}',
                estilo_firma
            ))
        else:
            elementos.append(Paragraph(
                f'<b>Entregó:</b> Secretaría Administrativa',
                estilo_firma
            ))
            elementos.append(Paragraph(
                f'<b>Recibió:</b> {solicitante}',
                estilo_firma
            ))

    doc.build(elementos)
    buffer.seek(0)
    return buffer


def generar_pdf_blanco(trabajador_nombre, tipo='pedido'):
    """
    Genera un PDF en blanco del formato institucional para que el trabajador pueda llenarlo a mano.
    tipo: 'pedido' → formato R-AP-33-01-01
          'salida' → formato R-AP-33-01-02
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Spacer, Paragraph)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'FormatoTitulo', parent=styles['Heading2'],
        alignment=TA_CENTER, fontSize=12, spaceAfter=4
    )
    estilo_subtitulo = ParagraphStyle(
        'FormatoSub', parent=styles['Normal'],
        alignment=TA_CENTER, fontSize=9, textColor=colors.grey,
        spaceAfter=12
    )
    estilo_campo = ParagraphStyle(
        'Campo', parent=styles['Normal'],
        fontSize=9, spaceAfter=2
    )
    estilo_firma = ParagraphStyle(
        'Firma', parent=styles['Normal'],
        fontSize=9, spaceBefore=20
    )

    elementos = []

    if tipo == 'pedido':
        titulo = 'REQUISICIÓN DE MATERIAL'
        codigo = 'R-AP-33-01-01 V4'
    else:
        titulo = 'SALIDA DE ALMACÉN'
        codigo = 'R-AP-33-01-02'

    elementos.append(Paragraph('SECRETARÍA ADMINISTRATIVA', estilo_titulo))
    elementos.append(Paragraph(titulo, estilo_titulo))
    elementos.append(Paragraph(codigo, estilo_subtitulo))

    fecha_str = datetime.utcnow().strftime('%d/%m/%Y')
    
    elementos.append(Paragraph(
        f'<b>Fecha:</b> {fecha_str} &nbsp;&nbsp; '
        f'<b>Folio:</b> ________________',
        estilo_campo
    ))
    elementos.append(Paragraph(
        f'<b>Departamento:</b> Mantenimiento',
        estilo_campo
    ))
    elementos.append(Spacer(1, 10))

    # Tabla de materiales en blanco
    encabezados = ['Cantidad', 'Unidad', 'Descripción']
    data = [encabezados]
    for _ in range(10):  # 10 filas en blanco
        data.append(['', '', ''])

    col_widths = [1.2*inch, 1.2*inch, 4*inch]
    tabla = Table(data, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C41E2A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F5F5F5')]),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elementos.append(tabla)

    if tipo == 'pedido':
        elementos.append(Paragraph(
            f'<b>Solicitó:</b> {trabajador_nombre}',
            estilo_firma
        ))
    else:
        elementos.append(Paragraph(
            f'<b>Entregó:</b> Secretaría Administrativa',
            estilo_firma
        ))
        elementos.append(Paragraph(
            f'<b>Recibió:</b> {trabajador_nombre}',
            estilo_firma
        ))

    doc.build(elementos)
    buffer.seek(0)
    return buffer


def generar_xlsx_pedido_trabajador(trabajador_nombre: str, items=None,
                                    area: str = None) -> BytesIO:
    """
    Genera el formato R-AP-33-01-01 (Requisición de Material) en Excel
    para un trabajador de mantenimiento con filas en blanco para llenado manual.

    trabajador_nombre: nombre que aparece en el campo "SOLICITÓ"
    area: área de trabajo del trabajador; aparece en el campo "DEPARTAMENTO"
    items: lista opcional de dicts {cantidad, unidad, descripcion}
    Retorna un BytesIO con el .xlsx.
    """
    if items is None:
        items = []

    departamento = area if area else 'Mantenimiento'

    wb = openpyxl.load_workbook(PLANTILLA_PEDIDO)
    ws = wb.active

    fecha = datetime.utcnow()

    # Bloque 1 (filas 1-19)
    _llenar_bloque_pedido(ws, 1, items, departamento, fecha, trabajador_nombre, 1)
    # Bloque 2 — copia (filas 20-38)
    _llenar_bloque_pedido(ws, 20, items, departamento, fecha, trabajador_nombre, 1)

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generar_xlsx_salida_trabajador(trabajador_nombre: str, items=None,
                                   area: str = None) -> BytesIO:
    """
    Genera el formato R-AP-33-01-02 (Salida de Almacén) en Excel
    para un trabajador de mantenimiento con filas en blanco para llenado manual.

    trabajador_nombre: nombre que aparece en el campo "RECIBIÓ"
    area: área de trabajo del trabajador; aparece en el campo "DEPARTAMENTO"
    items: lista opcional de dicts {cantidad, unidad, descripcion}
    Retorna un BytesIO con el .xlsx.
    """
    if items is None:
        items = []

    departamento = area if area else 'Mantenimiento'

    wb = openpyxl.load_workbook(PLANTILLA_SALIDA)
    ws = wb.active

    fecha = datetime.utcnow()

    # Bloque 1 (filas 1-19)
    _llenar_bloque_salida(ws, 1, items, departamento, fecha,
                          'Secretaría Administrativa', trabajador_nombre, 1)
    # Bloque 2 — copia (filas 20-38)
    _llenar_bloque_salida(ws, 20, items, departamento, fecha,
                          'Secretaría Administrativa', trabajador_nombre, 1)

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generar_xlsx_informe_general(trabajador_nombre: str, area: str = None,
                                  items=None) -> BytesIO:
    """
    Genera un libro Excel con DOS hojas:
      - "Entrada"  → Formato R-AP-33-01-01 (Requisición de Material)
      - "Salida"   → Formato R-AP-33-01-02 (Salida de Almacén)
    Ambas hojas pre-llenadas con el nombre, área e items del trabajador.
    Retorna un BytesIO con el .xlsx.
    """
    if items is None:
        items = []

    departamento = area if area else 'Mantenimiento'
    fecha = datetime.utcnow()

    # ── Hoja Entrada (R-AP-33-01-01) ─────────────────────────────────────────
    wb_entrada = openpyxl.load_workbook(PLANTILLA_PEDIDO)
    ws_entrada = wb_entrada.active
    ws_entrada.title = 'Entrada'
    _llenar_bloque_pedido(ws_entrada, 1,  items, departamento, fecha, trabajador_nombre, 1)
    _llenar_bloque_pedido(ws_entrada, 20, items, departamento, fecha, trabajador_nombre, 1)

    # ── Hoja Salida (R-AP-33-01-02) ──────────────────────────────────────────
    wb_salida = openpyxl.load_workbook(PLANTILLA_SALIDA)
    ws_salida = wb_salida.active
    ws_salida.title = 'Salida'
    _llenar_bloque_salida(ws_salida, 1,  items, departamento, fecha,
                          'Secretaría Administrativa', trabajador_nombre, 1)
    _llenar_bloque_salida(ws_salida, 20, items, departamento, fecha,
                          'Secretaría Administrativa', trabajador_nombre, 1)

    # ── Combinar en un solo libro con 2 hojas ────────────────────────────────
    wb_final = openpyxl.Workbook()
    wb_final.remove(wb_final.active)   # elimina la hoja vacía por defecto

    def _copiar_hoja(wb_origen, ws_origen, wb_destino, titulo):
        ws_dest = wb_destino.create_sheet(title=titulo)
        # Copiar dimensiones de columnas
        for col_letter, col_dim in ws_origen.column_dimensions.items():
            ws_dest.column_dimensions[col_letter].width = col_dim.width
        # Copiar filas con estilos
        for row in ws_origen.iter_rows():
            for cell in row:
                dest_cell = ws_dest.cell(row=cell.row, column=cell.column,
                                         value=cell.value)
                _copiar_estilo(cell, dest_cell)
        # Copiar merges
        for merged_range in ws_origen.merged_cells.ranges:
            ws_dest.merge_cells(str(merged_range))
        # Copiar alturas de filas
        for row_idx, row_dim in ws_origen.row_dimensions.items():
            ws_dest.row_dimensions[row_idx].height = row_dim.height

    _copiar_hoja(wb_entrada, ws_entrada, wb_final, 'Entrada')
    _copiar_hoja(wb_salida,  ws_salida,  wb_final, 'Salida')

    wb_entrada.close()
    wb_salida.close()

    buffer = BytesIO()
    wb_final.save(buffer)
    wb_final.close()
    buffer.seek(0)
    return buffer
